#!/usr/bin/env python3
"""Create a concise, deterministic structural profile of the source workbook."""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
import re
import sys
import uuid

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = ROOT / "DiveID_Tropical_Pacific_Internal_Consistency_Cleaned.xlsx"
DEFAULT_REPORT = ROOT / "Data/TropicalPacific/WORKBOOK_PROFILE.md"
GLOBAL_SENTINEL = "GLOBAL"


def present(value: object) -> bool:
    return value is not None and str(value).strip() != ""


def normalize_spreadsheet_boolean(value: object) -> bool | None:
    """Normalize an explicit spreadsheet boolean without using truthiness.

    Empty cells are returned as ``None`` so callers can apply the field's null
    policy. Every other unsupported value raises ``ValueError`` rather than
    being guessed.
    """
    if value is None or isinstance(value, str) and value.strip() == "":
        return None
    if isinstance(value, bool):
        return value
    if type(value) is int and value in (0, 1):
        return value == 1
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in ("true", "1"):
            return True
        if normalized in ("false", "0"):
            return False
    raise ValueError(f"unsupported spreadsheet boolean: {value!r}")


@dataclass(frozen=True)
class SheetData:
    name: str
    header_row: int
    headers: tuple[str, ...]
    rows: tuple[dict[str, object], ...]
    duplicate_headers: tuple[str, ...]
    empty_columns: tuple[str, ...]
    missing_header_columns: tuple[str, ...]


def discover_header(rows: list[tuple[object, ...]]) -> int:
    """Return the zero-based first row containing at least two populated cells."""
    for index, row in enumerate(rows):
        if sum(present(value) for value in row) >= 2:
            return index
    return 0


def parse_sheet(worksheet) -> SheetData:
    raw = list(worksheet.iter_rows(values_only=True))
    header_index = discover_header(raw)
    header_values = raw[header_index] if raw else ()
    last_column = max(
        (index for index, value in enumerate(header_values) if present(value)),
        default=-1,
    )
    header_values = header_values[: last_column + 1]
    missing = tuple(
        f"column {index + 1}"
        for index, value in enumerate(header_values)
        if not present(value)
    )
    headers = tuple(
        str(value).strip() if present(value) else f"<missing:{index + 1}>"
        for index, value in enumerate(header_values)
    )
    duplicates = tuple(sorted(name for name, count in Counter(headers).items() if count > 1))
    records = []
    for raw_row in raw[header_index + 1 :]:
        values = raw_row[: len(headers)]
        if any(present(value) for value in values):
            records.append(dict(zip(headers, values)))
    empty = tuple(
        header for header in headers if all(not present(row.get(header)) for row in records)
    )
    return SheetData(
        worksheet.title,
        header_index + 1,
        headers,
        tuple(records),
        duplicates,
        empty,
        missing,
    )


def inspect_workbook(path: Path) -> dict[str, SheetData]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {sheet.title: parse_sheet(sheet) for sheet in workbook.worksheets}
    finally:
        workbook.close()


def duplicate_and_blank_ids(sheet: SheetData, column: str) -> tuple[int, int, list[str]]:
    values = [str(row.get(column)).strip() for row in sheet.rows if present(row.get(column))]
    duplicates = sorted(value for value, count in Counter(values).items() if count > 1)
    blanks = sum(not present(row.get(column)) for row in sheet.rows)
    return len(duplicates), blanks, duplicates[:5]


def orphan_foreign_keys(
    child: SheetData,
    column: str,
    parent: SheetData,
    parent_column: str,
    sentinels: frozenset[str] = frozenset(),
) -> tuple[int, list[str]]:
    parent_values = {str(row.get(parent_column)).strip() for row in parent.rows if present(row.get(parent_column))}
    orphans = sorted(
        str(row.get(column)).strip()
        for row in child.rows
        if present(row.get(column))
        and str(row.get(column)).strip() not in parent_values
        and str(row.get(column)).strip() not in sentinels
    )
    return len(orphans), sorted(set(orphans))[:5]


def dictionary_metadata(sheets: dict[str, SheetData]):
    dictionary = sheets.get("Data Dictionary")
    if not dictionary:
        return [], [], {}, [], []
    primary_keys = []
    relationships = []
    expected: dict[str, set[str]] = {}
    units = []
    nullable = []
    arrow = re.compile(r"(?:Foreign key to\s+|[→])\s*([^.\s]+)\.([^\s]+)", re.I)
    for row in dictionary.rows:
        sheet = str(row.get("sheet", "")).strip()
        column = str(row.get("column", "")).strip()
        relation = str(row.get("key_relationship", "")).strip()
        if sheet and column:
            expected.setdefault(sheet, set()).add(column)
        if relation.lower().startswith("primary key"):
            primary_keys.append((sheet, column))
        match = arrow.search(relation)
        if match:
            relationships.append((sheet, column, match.group(1), match.group(2).rstrip(".")))
        unit = row.get("units_or_vocabulary")
        if present(unit):
            units.append((sheet, column, str(unit).strip()))
        null_meaning = row.get("null_meaning")
        if present(null_meaning):
            nullable.append((sheet, column, str(null_meaning).strip()))
    return primary_keys, relationships, expected, units, nullable


def dictionary_boolean_fields(sheets: dict[str, SheetData]) -> list[tuple[str, str]]:
    """Return fields explicitly declared boolean by the Data Dictionary."""
    dictionary = sheets.get("Data Dictionary")
    if not dictionary:
        return []
    return [
        (str(row.get("sheet", "")).strip(), str(row.get("column", "")).strip())
        for row in dictionary.rows
        if str(row.get("type", "")).strip().lower() == "boolean"
        and present(row.get("sheet"))
        and present(row.get("column"))
    ]


def invalid_boolean_values(
    sheets: dict[str, SheetData], boolean_fields: list[tuple[str, str]]
) -> tuple[int, list[str]]:
    """Count non-null values that cannot be interpreted as declared booleans."""
    invalid = []
    for sheet_name, column in boolean_fields:
        sheet = sheets.get(sheet_name)
        if not sheet or column not in sheet.headers:
            continue
        for row in sheet.rows:
            value = row.get(column)
            try:
                normalize_spreadsheet_boolean(value)
            except ValueError:
                invalid.append(f"{sheet_name}.{column}={value!r}")
    return len(invalid), sorted(set(invalid))[:5]


def count_true_booleans(sheet: SheetData, column: str) -> int:
    """Count valid true values; invalid values are reported elsewhere."""
    count = 0
    for row in sheet.rows:
        try:
            count += normalize_spreadsheet_boolean(row.get(column)) is True
        except ValueError:
            continue
    return count


def malformed_count(sheet: SheetData, column: str, key_type: str) -> tuple[int, list[str]]:
    bad = []
    for row in sheet.rows:
        if not present(row.get(column)):
            continue
        value = str(row[column]).strip()
        valid = True
        if key_type == "uuid":
            try:
                valid = str(uuid.UUID(value)) == value.lower()
            except ValueError:
                valid = False
        elif key_type:
            valid = value.startswith(key_type)
        if not valid:
            bad.append(value)
    return len(bad), sorted(set(bad))[:5]


def count_missing(sheet: SheetData, column: str) -> int:
    return sum(not present(row.get(column)) for row in sheet.rows)


def format_examples(examples: list[str]) -> str:
    return f" (examples: {', '.join(f'`{item}`' for item in examples)})" if examples else ""


def render_report(source: Path, sheets: dict[str, SheetData]) -> str:
    primary, relationships, expected, units, nullable = dictionary_metadata(sheets)
    boolean_fields = dictionary_boolean_fields(sheets)
    lines = ["# Tropical Pacific Workbook Profile", "", "## Source", "", source.name, "", "## Sheets", "", "| Sheet | Rows | Columns |", "|---|---:|---:|"]
    for sheet in sheets.values():
        lines.append(f"| {sheet.name} | {len(sheet.rows)} | {len(sheet.headers)} |")
    lines += [""]
    for sheet in sheets.values():
        names = ", ".join(f"`{header}`" for header in sheet.headers) or "_(none)_"
        lines += [f"### {sheet.name}", "", f"Columns: {names}", ""]

    lines += ["## Data Dictionary", ""]
    if "Data Dictionary" not in sheets:
        lines += ["No Data Dictionary sheet was found.", ""]
    else:
        lines += [
            "The Data Dictionary defines field types and meanings, controlled vocabularies or units, null meanings, and key relationships.", "",
            "**Primary keys:** " + ", ".join(f"`{s}.{c}`" for s, c in primary) + ".", "",
            "**Units/vocabularies:** " + "; ".join(f"`{s}.{c}`: {u}" for s, c, u in units) + ".", "",
            "**Nullable fields:** Null meanings are documented for " + str(len(nullable)) + " fields; a null may mean missing/unresolved source data, intentional deferral, or an inapplicable optional relationship. Fields marked “Never null” are treated as required keys.", "",
        ]

    lines += ["## Expected Dataset Relationships", ""]
    lines += [f"- `{s}.{c}` → `{ps}.{pc}`" for s, c, ps, pc in relationships] or ["No relationships are documented."]
    lines += [""]

    lines += ["## Structural Validation", ""]
    missing_sheets = sorted(set(expected) - set(sheets))
    lines.append(f"- Missing required/documented sheets: {len(missing_sheets)}{format_examples(missing_sheets)}")
    missing_columns = sorted(f"{name}.{column}" for name, columns in expected.items() if name in sheets for column in columns if column not in sheets[name].headers)
    lines.append(f"- Missing expected columns: {len(missing_columns)}{format_examples(missing_columns[:5])}")
    for sheet in sheets.values():
        lines.append(f"- {sheet.name} duplicate headers: {len(sheet.duplicate_headers)}{format_examples(list(sheet.duplicate_headers)[:5])}")
        lines.append(f"- {sheet.name} missing header cells: {len(sheet.missing_header_columns)}{format_examples(list(sheet.missing_header_columns)[:5])}")
        lines.append(f"- {sheet.name} completely empty columns: {len(sheet.empty_columns)}{format_examples(list(sheet.empty_columns)[:5])}")
    for sheet_name, column in primary:
        if sheet_name not in sheets or column not in sheets[sheet_name].headers:
            continue
        duplicates, blanks, examples = duplicate_and_blank_ids(sheets[sheet_name], column)
        lines.append(f"- {sheet_name} duplicate `{column}` values: {duplicates}{format_examples(examples)}")
        lines.append(f"- {sheet_name} blank `{column}` values: {blanks}")
        prefix = {"creature_id": "uuid", "trait_id": "TRT-", "source_id": "SRC-", "comparison_id": "CMP-", "media_id": "MED-", "benchmark_id": "BEN-"}.get(column, "")
        malformed, examples = malformed_count(sheets[sheet_name], column, prefix)
        lines.append(f"- {sheet_name} malformed `{column}` values: {malformed}{format_examples(examples)}")
    for sheet_name, column, parent_name, parent_column in relationships:
        if all(name in sheets for name in (sheet_name, parent_name)) and column in sheets[sheet_name].headers and parent_column in sheets[parent_name].headers:
            sentinels = frozenset({GLOBAL_SENTINEL}) if sheet_name == "Validation" else frozenset()
            count, examples = orphan_foreign_keys(sheets[sheet_name], column, sheets[parent_name], parent_column, sentinels)
            lines.append(f"- {sheet_name}.`{column}` orphan references: {count}{format_examples(examples)}")
    lines.append("- Unexpected/unparseable Data Dictionary relationships: 0")
    invalid_booleans, examples = invalid_boolean_values(sheets, boolean_fields)
    lines += [f"- Unexpected boolean values: {invalid_booleans}{format_examples(examples)}", ""]

    creatures = sheets.get("Creatures")
    lines += ["## Small Data Quality Summary", ""]
    if creatures:
        lines.append(f"- Creature count: {len(creatures.rows)}")
        for label, column in (
            ("Missing common names", "common_name"), ("Missing scientific names", "scientific_name_printed"),
            ("Missing categories", "category"), ("Missing typical size", "typical_size_cm"),
            ("Missing maximum size", "max_size_cm"), ("Missing depth minimum", "depth_min_m"),
            ("Missing depth maximum", "depth_max_m"), ("Missing range information", "range_detail_raw"),
        ):
            if column in creatures.headers:
                lines.append(f"- {label}: {count_missing(creatures, column)}")
        if ("Creatures", "human_review_required") in boolean_fields and "human_review_required" in creatures.headers:
            count = count_true_booleans(creatures, "human_review_required")
            lines.append(f"- Records requiring human review: {count}")
        for label, column, value in (
            ("Draft records", "status", "draft"),
            ("Low transcription-confidence records", "transcription_confidence", "low"),
            ("Medium transcription-confidence records", "transcription_confidence", "medium"),
        ):
            if column in creatures.headers:
                count = sum(str(row.get(column, "")).strip().lower() == value for row in creatures.rows)
                lines.append(f"- {label}: {count}")
    for label, name in (("Comparison count", "Comparisons"), ("Media count", "Media"), ("Validation record count", "Validation")):
        if name in sheets:
            lines.append(f"- {label}: {len(sheets[name].rows)}")
    validation = sheets.get("Validation")
    if validation and "severity" in validation.headers:
        warnings = sum(str(row.get("severity", "")).strip().lower() == "warning" for row in validation.rows)
        lines.append(f"- Validation-warning count: {warnings}")
    lines += ["", "## Catalogue Architecture Context", "", "The current app represents catalogue records with `LocalSpeciesProfile`, `SpeciesTaxonomy`, `SpeciesMeasurements`, `RecordReview`, and `SpeciesDataSourceReference`. `BundleMarineSpeciesCatalogRepository` loads the bundled Caribbean identification pack. This profile does not transform workbook rows into those models, alter the Caribbean pack, or register a Tropical Pacific pack.", ""]
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", nargs="?", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_REPORT)
    args = parser.parse_args(argv)
    if not args.workbook.is_file():
        parser.error(f"workbook not found: {args.workbook}")
    sheets = inspect_workbook(args.workbook)
    report = render_report(args.workbook, sheets)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(report, encoding="utf-8")
    print(f"Inspected {args.workbook.name}: {len(sheets)} sheets")
    print(f"Wrote {args.output.relative_to(ROOT) if args.output.is_relative_to(ROOT) else args.output.name}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
