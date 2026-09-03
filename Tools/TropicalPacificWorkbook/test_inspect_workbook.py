from collections import Counter
import hashlib
from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from Tools.TropicalPacificWorkbook.inspect_workbook import (
    DEFAULT_REPORT,
    DEFAULT_WORKBOOK,
    count_true_booleans,
    dictionary_boolean_fields,
    duplicate_and_blank_ids,
    inspect_workbook,
    normalize_spreadsheet_boolean,
    orphan_foreign_keys,
    render_report,
)


class WorkbookInspectorTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.path = Path(self.temporary_directory.name) / "tiny.xlsx"
        workbook = Workbook()
        creatures = workbook.active
        creatures.title = "Creatures"
        creatures.append(["creature_id", "common_name"])
        creatures.append(["11111111-1111-1111-1111-111111111111", "One"])
        creatures.append(["11111111-1111-1111-1111-111111111111", "Duplicate"])
        creatures.append([None, "Blank"])
        traits = workbook.create_sheet("Traits")
        traits.append(["trait_id", "creature_id"])
        traits.append(["TRT-1", "11111111-1111-1111-1111-111111111111"])
        traits.append(["TRT-2", "22222222-2222-2222-2222-222222222222"])
        validation = workbook.create_sheet("Validation")
        validation.append(["entity_id", "issue"])
        validation.append(["GLOBAL", "Workbook note"])
        workbook.save(self.path)

    def tearDown(self):
        self.temporary_directory.cleanup()

    def test_discovery_headers_and_row_count(self):
        sheets = inspect_workbook(self.path)
        self.assertEqual(list(sheets), ["Creatures", "Traits", "Validation"])
        self.assertEqual(sheets["Creatures"].headers, ("creature_id", "common_name"))
        self.assertEqual(len(sheets["Creatures"].rows), 3)

    def test_duplicate_and_blank_primary_ids(self):
        creatures = inspect_workbook(self.path)["Creatures"]
        duplicates, blanks, _ = duplicate_and_blank_ids(creatures, "creature_id")
        self.assertEqual((duplicates, blanks), (1, 1))

    def test_orphan_foreign_keys(self):
        sheets = inspect_workbook(self.path)
        count, examples = orphan_foreign_keys(sheets["Traits"], "creature_id", sheets["Creatures"], "creature_id")
        self.assertEqual(count, 1)
        self.assertEqual(examples, ["22222222-2222-2222-2222-222222222222"])

    def test_global_sentinel_is_not_an_orphan(self):
        sheets = inspect_workbook(self.path)
        count, _ = orphan_foreign_keys(sheets["Validation"], "entity_id", sheets["Creatures"], "creature_id", frozenset({"GLOBAL"}))
        self.assertEqual(count, 0)

    def test_read_only_inspection_does_not_modify_workbook(self):
        before = hashlib.sha256(self.path.read_bytes()).digest()
        inspect_workbook(self.path)
        after = hashlib.sha256(self.path.read_bytes()).digest()
        self.assertEqual(before, after)
        workbook = load_workbook(self.path, read_only=True)
        self.assertTrue(workbook.read_only)
        workbook.close()

    def test_supported_boolean_representations(self):
        cases = (
            (True, True),
            (False, False),
            (1, True),
            (0, False),
            ("1", True),
            ("0", False),
            ("TRUE", True),
            ("FALSE", False),
            ("true", True),
            ("false", False),
        )
        for source, expected in cases:
            with self.subTest(source=source):
                self.assertIs(normalize_spreadsheet_boolean(source), expected)

    def test_committed_workbook_boolean_representation_and_count(self):
        sheets = inspect_workbook(DEFAULT_WORKBOOK)
        creatures = sheets["Creatures"]
        values = [row["human_review_required"] for row in creatures.rows]

        self.assertEqual(
            dictionary_boolean_fields(sheets),
            [("Creatures", "human_review_required")],
        )
        self.assertEqual(
            Counter((type(value), value) for value in values),
            {(bool, True): 1650},
        )
        self.assertEqual(count_true_booleans(creatures, "human_review_required"), 1650)

    def test_unexpected_boolean_is_reported(self):
        for value in (2, -1, 0.5, "yes", "no", object()):
            with self.subTest(value=value):
                with self.assertRaises(ValueError):
                    normalize_spreadsheet_boolean(value)

        workbook = Workbook()
        creatures = workbook.active
        creatures.title = "Creatures"
        creatures.append(["human_review_required"])
        creatures.append(["yes"])
        dictionary = workbook.create_sheet("Data Dictionary")
        dictionary.append(["sheet", "column", "type"])
        dictionary.append(["Creatures", "human_review_required", "boolean"])
        workbook.save(self.path)

        report = render_report(self.path, inspect_workbook(self.path))

        self.assertIn("- Unexpected boolean values: 1", report)
        self.assertIn("`Creatures.human_review_required='yes'`", report)

    def test_committed_workbook_reproduces_complete_report(self):
        generated = render_report(DEFAULT_WORKBOOK, inspect_workbook(DEFAULT_WORKBOOK))
        self.assertEqual(generated, DEFAULT_REPORT.read_text(encoding="utf-8"))


if __name__ == "__main__":
    unittest.main()
